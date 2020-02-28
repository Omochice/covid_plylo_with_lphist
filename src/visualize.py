import csv
import json
import os
import string
import sys
from itertools import product
from pathlib import Path

import openpyxl
from openpyxl.styles.alignment import Alignment

project_dir = Path(__file__).resolve().parents[2]


def make_codon(data: dict,
               filename: str,
               save_dir: str,
               axis_names=None,
               save_xlxs=True,
               save_csv=True) -> None:
    """コドン表みたいなexcelをつくる"""
    base_dir = os.path.dirname(os.path.abspath(__file__))

    book = openpyxl.Workbook()
    book.create_sheet()

    sheet = book[book.get_sheet_names()[0]]

    if axis_names is None:
        keys = data.keys()
        axis_names = list(set("".join(keys)))

    axis_lenght = len(axis_names)

    "1stの列を作る"
    sheet["A1"] = "1st"
    sheet.merge_cells("A1:A2")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for i, char in enumerate(axis_names):
        start_cell = str(3 + (axis_lenght * i))
        lase_cell = str(2 + (axis_lenght * (i + 1)))
        cell_name = "A" + start_cell
        sheet[cell_name] = char
        sheet.merge_cells(f"{cell_name}:A{lase_cell}")
        sheet[cell_name].alignment = Alignment(horizontal="center",
                                               vertical="center")

    "2ndの列を作る"
    alphabet = string.ascii_uppercase
    sheet["B1"] = "2nd"
    sheet.merge_cells(f"B1:{alphabet[axis_lenght]}1")
    sheet["B1"].alignment = Alignment(horizontal="center", vertical="center")
    for i, char in enumerate(axis_names):
        cell_name = alphabet[i + 1] + "2"
        sheet[cell_name] = char

    "3rdの列を作る"
    collunm_3rd = alphabet[axis_lenght + 1]
    third_name = f"{collunm_3rd}1"
    sheet[third_name] = "3rd"
    sheet.merge_cells(f"{third_name}:{collunm_3rd}2")
    sheet[third_name].alignment = Alignment(horizontal="center",
                                            vertical="center")
    for j in range(axis_lenght):
        for i, char in enumerate(axis_names):
            row = str(3 + (axis_lenght * j) + i)
            sheet[collunm_3rd + row] = char

    "実データで埋めていく"
    use_collonms = alphabet[1:axis_lenght + 1]
    use_rows = map(lambda x: x + 3, range(axis_lenght**2))
    for row, collunm in product(use_rows, use_collonms):
        cell_name = f"{collunm}{row}"
        for_first = (row - 3) // axis_lenght * axis_lenght + 3
        first = sheet[f"A{for_first}"].value
        second = sheet[f"{collunm}2"].value
        third = sheet[f"{collunm_3rd}{row}"].value
        key = first + second + third
        sheet[cell_name] = data.get(key, None)

    if save_xlxs:
        book.save(os.path.join(save_dir, f"{filename}.xlsx"))

    if save_csv:
        with open(os.path.join(save_dir, f"{filename}.csv"), "w") as f:
            writer = csv.writer(f)
            for row in sheet.rows:
                writer.writerow([cell.value for cell in row])


if __name__ == "__main__":
    with open(sys.argv[1], "r") as f:
        data = json.load(f)
    axis = ["a", "t", "g", "c"]

    filename = os.path.splitext(os.path.basename(sys.argv[1]))[0]
    save_dir = os.path.join(project_dir, "data", "interim", "weight")
    os.makedirs(save_dir, exist_ok=True)
    make_codon(data=data,
               filename=filename,
               save_dir=save_dir,
               axis_names=axis)
